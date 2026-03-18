from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify
from flask_login import login_required, current_user
from app import db
from app.models import Drawing, Photo, Delivery
from datetime import date, datetime
import os, uuid, io
from werkzeug.utils import secure_filename
import openpyxl

drawing = Blueprint("drawing", __name__)

ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "gif"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ─── 図面一覧 ──────────────────────────────────────
@drawing.route("/")
@login_required
def drawing_list():
    drawings = Drawing.query.filter_by(is_deleted=False).order_by(
        Drawing.project_no.asc(),
        Drawing.drawing_no.asc()
    ).all()
    return render_template("drawing_list.html", drawings=drawings)


# ─── 新規登録 ──────────────────────────────────────
@drawing.route("/drawing/new", methods=["GET", "POST"])
@login_required
def drawing_new():
    if request.method == "POST":
        project_no = request.form.get("project_no", "").strip()
        drawing_no = request.form.get("drawing_no", "").strip()
        vendor     = request.form.get("vendor", "").strip()
        order_date_str = request.form.get("order_date", "").strip()
        due_date_str   = request.form.get("due_date", "").strip()

        error = None
        order_date = None
        due_date   = None

        try:
            order_date = datetime.strptime(order_date_str, "%Y/%m/%d").date()
        except ValueError:
            error = "発注日は YYYY/MM/DD 形式で入力してください（例：2026/02/10）"

        if not error:
            try:
                due_date = datetime.strptime(due_date_str, "%Y/%m/%d").date()
            except ValueError:
                error = "納期は YYYY/MM/DD 形式で入力してください（例：2026/02/10）"

        if error:
            form_data = {
                "project_no": project_no,
                "drawing_no": drawing_no,
                "vendor":     vendor,
                "order_date": order_date_str,
                "due_date":   due_date_str,
            }
            return render_template("drawing_new.html", error=error, form_data=form_data)

        drawing_obj = Drawing(
            project_no=project_no,
            drawing_no=drawing_no,
            vendor=vendor,
            order_date=order_date,
            due_date=due_date,
        )
        db.session.add(drawing_obj)
        db.session.commit()
        return redirect(url_for("drawing.drawing_list"))

    return render_template("drawing_new.html")


# ─── 図面の論理削除 ────────────────────────────────
@drawing.route("/drawing/delete", methods=["POST"])
@login_required
def drawing_delete():
    data = request.get_json()
    delete_ids = [int(i) for i in data.get("ids", [])]
    Drawing.query.filter(Drawing.id.in_(delete_ids)).update(
        {"is_deleted": True}, synchronize_session=False
    )
    db.session.commit()
    return jsonify({"ok": True})


# ─── Excel取り込み ────────────────────────────────
@drawing.route("/drawing/import", methods=["GET", "POST"])
@login_required
def drawing_import():
    COLUMN_MAP = {
        "工事番号": "project_no",
        "図面番号": "drawing_no",
        "発注先":   "vendor",
        "発注日":   "order_date",
        "納期":     "due_date",
    }

    def parse_date(val):
        """セルの値を date 型に変換。失敗時は None を返す。"""
        if val is None:
            return None
        if isinstance(val, (date, datetime)):
            return val.date() if isinstance(val, datetime) else val
        s = str(val).strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    # ── GET：通常表示 ──────────────────────────────
    if request.method == "GET":
        return render_template("drawing_import.html")

    # ── POST：上書き確定（確認画面からの送信） ──────
    if request.form.get("action") == "overwrite_confirm":
        try:
            pending = session.pop("import_pending", [])
            overwrite_ids = set(request.form.getlist("overwrite_ids"))
            overwritten = 0
            for item in pending:
                if str(item["drawing_id"]) not in overwrite_ids:
                    continue
                d = Drawing.query.get(item["drawing_id"])
                if d:
                    d.project_no = item["project_no"]
                    d.drawing_no = item["drawing_no"]
                    d.vendor     = item["vendor"]
                    if item["order_date"]:
                        d.order_date = date.fromisoformat(item["order_date"])
                    if item["due_date"]:
                        d.due_date = date.fromisoformat(item["due_date"])
                    overwritten += 1
            db.session.commit()
            imported_count = session.pop("import_imported", 0)
            skipped_count  = session.pop("import_skipped", 0)
            return render_template("drawing_import.html",
                                   result={
                                       "imported":    imported_count,
                                       "skipped":     skipped_count,
                                       "overwritten": overwritten,
                                   })
        except Exception as e:
            db.session.rollback()
            return render_template("drawing_import.html",
                                   error=f"上書き処理中にエラーが発生しました: {str(e)}")

    # ── POST：Excelファイルの新規読み込み ───────────
    file = request.files.get("excel_file")
    if not file or not file.filename.endswith((".xlsx", ".xls")):
        return render_template("drawing_import.html",
                               error="Excelファイル（.xlsx / .xls）を選択してください。")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active

        headers = [str(cell.value).strip() if cell.value is not None else ""
                   for cell in ws[1]]

        col_index = {}
        for excel_col, model_attr in COLUMN_MAP.items():
            if excel_col in headers:
                col_index[model_attr] = headers.index(excel_col)

        if "project_no" not in col_index and "drawing_no" not in col_index:
            return render_template("drawing_import.html",
                                   error="「工事番号」「図面番号」いずれのヘッダーも見つかりませんでした。"
                                         "1行目にヘッダーが記載されているか確認してください。")

        imported  = 0
        skipped   = 0
        pending   = []   # 上書き候補（納入日未登録の既存レコード）

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            def cell(attr):
                idx = col_index.get(attr)
                if idx is None:
                    return None
                v = row[idx]
                return str(v).strip() if v is not None else ""

            project_no = cell("project_no") or ""
            drawing_no = cell("drawing_no") or ""

            if not project_no and not drawing_no:
                skipped += 1
                continue

            vendor     = cell("vendor") or ""
            order_date = parse_date(row[col_index["order_date"]]) \
                         if "order_date" in col_index else date.today()
            due_date   = parse_date(row[col_index["due_date"]]) \
                         if "due_date" in col_index else date.today()

            exists = Drawing.query.filter_by(
                project_no=project_no,
                drawing_no=drawing_no,
                is_deleted=False,
            ).first()

            if exists:
                # 納入日が登録済み → 完全スキップ
                has_delivery = (
                    exists.delivery is not None
                    and exists.delivery.delivered_at is not None
                )
                if has_delivery:
                    skipped += 1
                else:
                    # 納入日未登録 → 上書き候補に追加
                    pending.append({
                        "drawing_id": exists.id,
                        "project_no": project_no,
                        "drawing_no": drawing_no,
                        "vendor":     vendor,
                        "order_date": order_date.isoformat() if order_date else None,
                        "due_date":   due_date.isoformat()   if due_date   else None,
                    })
                continue

            # 新規登録
            drawing_obj = Drawing(
                project_no = project_no,
                drawing_no = drawing_no,
                vendor     = vendor,
                order_date = order_date or date.today(),
                due_date   = due_date   or date.today(),
                created_by = current_user.id,
            )
            db.session.add(drawing_obj)
            imported += 1

        db.session.commit()

        # 上書き候補がある → 確認画面へ
        if pending:
            session["import_pending"]  = pending
            session["import_imported"] = imported
            session["import_skipped"]  = skipped
            return render_template("drawing_import.html",
                                   pending=pending,
                                   imported=imported,
                                   skipped=skipped)

        return render_template("drawing_import.html",
                               result={"imported": imported, "skipped": skipped,
                                       "overwritten": 0})

    except Exception as e:
        db.session.rollback()
        return render_template("drawing_import.html",
                               error=f"取り込み中にエラーが発生しました: {str(e)}")


# ─── 納品登録（カートに追加） ──────────────────────
@drawing.route("/delivery/add/<int:drawing_id>", methods=["POST"])
@login_required
def delivery_add(drawing_id):
    cart = session.get("delivery_cart", [])
    if drawing_id not in cart:
        cart.append(drawing_id)
    session["delivery_cart"] = cart
    return redirect(url_for("drawing.drawing_list"))


# ─── 納品登録取り消し（カートから削除） ───────────
@drawing.route("/delivery/remove", methods=["POST"])
@login_required
def delivery_remove():
    data = request.get_json()
    remove_ids = [int(i) for i in data.get("ids", [])]
    cart = session.get("delivery_cart", [])
    cart = [i for i in cart if i not in remove_ids]
    session["delivery_cart"] = cart
    return jsonify({"ok": True, "cart": cart})


# ─── 写真アップロード画面 ──────────────────────────
@drawing.route("/photo/upload", methods=["GET", "POST"])
@login_required
def photo_upload():
    cart = session.get("delivery_cart", [])
    drawings = Drawing.query.filter(Drawing.id.in_(cart)).all() if cart else []

    if request.method == "POST":
        file = request.files.get("photo")
        drawing_ids = request.form.getlist("drawing_ids")

        if not file or not allowed_file(file.filename):
            return render_template("photo_upload.html", drawings=drawings,
                                   error="写真ファイルを選択してください。")

        ext = file.filename.rsplit(".", 1)[1].lower()
        saved_name = f"{uuid.uuid4().hex}.{ext}"
        from flask import current_app
        save_path = os.path.join(current_app.config["UPLOAD_FOLDER"], saved_name)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        file.save(save_path)

        photo = Photo(
            filename=saved_name,
            original_filename=secure_filename(file.filename),
            filepath=f"static/uploads/{saved_name}",
        )
        db.session.add(photo)
        db.session.flush()

        today = date.today()
        for did in drawing_ids:
            delivery = Delivery(
                drawing_id=int(did),
                photo_id=photo.id,
                delivered_at=today,
            )
            db.session.add(delivery)

        db.session.commit()
        session.pop("delivery_cart", None)
        return redirect(url_for("drawing.drawing_list"))

    return render_template("photo_upload.html", drawings=drawings)


# ─── 写真モーダル用API ─────────────────────────────
@drawing.route("/photo/<int:drawing_id>")
@login_required
def photo_data(drawing_id):
    delivery = Delivery.query.filter_by(drawing_id=drawing_id).first_or_404()
    return jsonify({
        "photo_url": url_for("static", filename=f"uploads/{delivery.photo.filename}"),
        "delivered_at": delivery.delivered_at.strftime("%Y/%m/%d"),
    })
