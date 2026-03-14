from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify
from flask_login import login_required, current_user
from app import db
from app.models import Drawing, Photo, Delivery, Vendor
from datetime import date, datetime
import os, uuid, io
from werkzeug.utils import secure_filename
import openpyxl

main = Blueprint("main", __name__)

ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "gif"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ─── 図面一覧 ──────────────────────────────────────
@main.route("/")
@login_required
def drawing_list():
    drawings = Drawing.query.filter_by(is_deleted=False).order_by(
        Drawing.project_no.asc(),
        Drawing.drawing_no.asc()
    ).all()
    return render_template("drawing_list.html", drawings=drawings)


# ─── 新規登録 ──────────────────────────────────────
@main.route("/drawing/new", methods=["GET", "POST"])
@login_required
def drawing_new():
    if request.method == "POST":
        # 入力値を取得
        project_no = request.form.get("project_no", "").strip()
        drawing_no = request.form.get("drawing_no", "").strip()
        vendor     = request.form.get("vendor", "").strip()
        order_date_str = request.form.get("order_date", "").strip()
        due_date_str   = request.form.get("due_date", "").strip()

        # バリデーション
        error = None
        order_date = None
        due_date   = None

        try:
            order_date = datetime.strptime(order_date_str, "%Y/%m/%d").date()
        except ValueError:
            error = "発注日は YYYY/MM/DD 形式で入力してください（例：2026/02/10）"

        if not error:
            try:
                due_date = datetime.strptime(due_date_str, "%Y/%m/%d").date()
            except ValueError:
                error = "納期は YYYY/MM/DD 形式で入力してください（例：2026/02/10）"

        # エラーがあれば入力値を保持してフォームを再表示
        if error:
            form_data = {
                "project_no": project_no,
                "drawing_no": drawing_no,
                "vendor":     vendor,
                "order_date": order_date_str,
                "due_date":   due_date_str,
            }
            return render_template("drawing_new.html", error=error, form_data=form_data)

        # DBに保存
        drawing = Drawing(
            project_no=project_no,
            drawing_no=drawing_no,
            vendor=vendor,
            order_date=order_date,
            due_date=due_date,
        )
        db.session.add(drawing)
        db.session.commit()
        return redirect(url_for("main.drawing_list"))

    return render_template("drawing_new.html")


# ─── 図面の論理削除 ────────────────────────────────
@main.route("/drawing/delete", methods=["POST"])
@login_required
def drawing_delete():
    """チェックされた drawing_id を論理削除する"""
    data = request.get_json()
    delete_ids = [int(i) for i in data.get("ids", [])]
    Drawing.query.filter(Drawing.id.in_(delete_ids)).update(
        {"is_deleted": True}, synchronize_session=False
    )
    db.session.commit()
    return jsonify({"ok": True})


# ─── Excel取り込み ────────────────────────────────
@main.route("/drawing/import", methods=["GET", "POST"])
@login_required
def drawing_import():
    """ExcelファイルからDrawingを一括登録する"""

    # ヘッダー行と取り込みカラムの定義
    # キー: Excelヘッダー名, 値: Drawingの属性名
    COLUMN_MAP = {
        "工事番号": "project_no",
        "図面番号": "drawing_no",
    }
    # Drawing登録に必須だが Excel にない項目のデフォルト値
    DEFAULTS = {
        "vendor":     "",
        "order_date": date.today(),
        "due_date":   date.today(),
    }

    if request.method == "GET":
        return render_template("drawing_import.html")

    # ── POST: ファイルを受け取って処理 ──────────────
    file = request.files.get("excel_file")
    if not file or not file.filename.endswith((".xlsx", ".xls")):
        return render_template("drawing_import.html",
                               error="Excelファイル（.xlsx / .xls）を選択してください。")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active  # 先頭シートを対象

        # 1行目をヘッダーとして読み込む
        headers = [str(cell.value).strip() if cell.value is not None else ""
                   for cell in ws[1]]

        # COLUMN_MAP のうちファイルに存在するものだけ使う
        col_index = {}  # {"project_no": 0, "drawing_no": 1, ...}
        for excel_col, model_attr in COLUMN_MAP.items():
            if excel_col in headers:
                col_index[model_attr] = headers.index(excel_col)

        if not col_index:
            return render_template("drawing_import.html",
                                   error="「工事番号」「図面番号」いずれのヘッダーも見つかりませんでした。"
                                         "1行目にヘッダーが記載されているか確認してください。")

        imported = 0
        skipped  = 0
        errors   = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 全セルが空の行はスキップ
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            project_no = str(row[col_index["project_no"]]).strip() \
                         if "project_no" in col_index else ""
            drawing_no = str(row[col_index["drawing_no"]]).strip() \
                         if "drawing_no" in col_index else ""

            # どちらも空ならスキップ
            if not project_no and not drawing_no:
                skipped += 1
                continue

            # 既に同じ project_no + drawing_no が存在すればスキップ
            exists = Drawing.query.filter_by(
                project_no=project_no,
                drawing_no=drawing_no,
                is_deleted=False,
            ).first()
            if exists:
                skipped += 1
                continue

            drawing = Drawing(
                project_no = project_no,
                drawing_no = drawing_no,
                vendor     = DEFAULTS["vendor"],
                order_date = DEFAULTS["order_date"],
                due_date   = DEFAULTS["due_date"],
                created_by = current_user.id,
            )
            db.session.add(drawing)
            imported += 1

        db.session.commit()

        return render_template("drawing_import.html",
                               result={"imported": imported, "skipped": skipped})

    except Exception as e:
        db.session.rollback()
        return render_template("drawing_import.html",
                               error=f"取り込み中にエラーが発生しました: {str(e)}")


# ─── 納品登録（カートに追加） ──────────────────────
@main.route("/delivery/add/<int:drawing_id>", methods=["POST"])
@login_required
def delivery_add(drawing_id):
    """納品登録ボタンを押した drawing_id をセッションに追加する"""
    cart = session.get("delivery_cart", [])
    if drawing_id not in cart:
        cart.append(drawing_id)
    session["delivery_cart"] = cart
    return redirect(url_for("main.drawing_list"))


# ─── 納品登録取り消し（カートから削除） ───────────
@main.route("/delivery/remove", methods=["POST"])
@login_required
def delivery_remove():
    """写真アップロード画面でチェックした drawing_id をセッションから削除する"""
    data = request.get_json()
    remove_ids = [int(i) for i in data.get("ids", [])]
    cart = session.get("delivery_cart", [])
    cart = [i for i in cart if i not in remove_ids]
    session["delivery_cart"] = cart
    return jsonify({"ok": True, "cart": cart})


# ─── 写真アップロード画面 ──────────────────────────
@main.route("/photo/upload", methods=["GET", "POST"])
@login_required
def photo_upload():
    cart = session.get("delivery_cart", [])
    drawings = Drawing.query.filter(Drawing.id.in_(cart)).all() if cart else []

    if request.method == "POST":
        file = request.files.get("photo")
        drawing_ids = request.form.getlist("drawing_ids")

        if not file or not allowed_file(file.filename):
            return render_template("photo_upload.html", drawings=drawings, error="写真ファイルを選択してください。")

        # ファイル保存
        ext = file.filename.rsplit(".", 1)[1].lower()
        saved_name = f"{uuid.uuid4().hex}.{ext}"
        from flask import current_app
        save_path = os.path.join(current_app.config["UPLOAD_FOLDER"], saved_name)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        file.save(save_path)

        # photos テーブルに登録
        photo = Photo(
            filename=saved_name,
            original_filename=secure_filename(file.filename),
            filepath=f"static/uploads/{saved_name}",
        )
        db.session.add(photo)
        db.session.flush()

        # deliveries テーブルに一括登録
        today = date.today()
        for did in drawing_ids:
            delivery = Delivery(
                drawing_id=int(did),
                photo_id=photo.id,
                delivered_at=today,
            )
            db.session.add(delivery)

        db.session.commit()

        # カートをクリア
        session.pop("delivery_cart", None)

        return redirect(url_for("main.drawing_list"))

    return render_template("photo_upload.html", drawings=drawings)


# ─── 写真モーダル用API ─────────────────────────────
@main.route("/photo/<int:drawing_id>")
@login_required
def photo_data(drawing_id):
    """納入日リンクをクリックしたときに写真URLをJSONで返す"""
    delivery = Delivery.query.filter_by(drawing_id=drawing_id).first_or_404()
    return jsonify({
        "photo_url": url_for("static", filename=f"uploads/{delivery.photo.filename}"),
        "delivered_at": delivery.delivered_at.strftime("%Y/%m/%d"),
    })


# ─── 発注先マスタ一覧 ──────────────────────────────
@main.route("/vendor")
@login_required
def vendor_list():
    vendors = Vendor.query.filter_by(is_deleted=False).order_by(
        Vendor.short_name.asc()
    ).all()
    return render_template("vendor_list.html", vendors=vendors)


# ─── 発注先マスタ新規登録 ──────────────────────────
@main.route("/vendor/new", methods=["GET", "POST"])
@login_required
def vendor_new():
    if request.method == "POST":
        name       = request.form.get("name", "").strip()
        short_name = request.form.get("short_name", "").strip()

        error = None
        if not name:
            error = "会社名を入力してください。"
        elif not short_name:
            error = "略称を入力してください。"

        if error:
            form_data = {"name": name, "short_name": short_name}
            return render_template("vendor_new.html", error=error, form_data=form_data)

        vendor = Vendor(
            name=name,
            short_name=short_name,
            created_by=current_user.id,
        )
        db.session.add(vendor)
        db.session.commit()
        return redirect(url_for("main.vendor_list"))

    return render_template("vendor_new.html")


# ─── 発注先マスタ論理削除 ──────────────────────────
@main.route("/vendor/delete", methods=["POST"])
@login_required
def vendor_delete():
    """チェックされた vendor_id を論理削除する"""
    data = request.get_json()
    delete_ids = [int(i) for i in data.get("ids", [])]
    Vendor.query.filter(Vendor.id.in_(delete_ids)).update(
        {"is_deleted": True}, synchronize_session=False
    )
    db.session.commit()
    return jsonify({"ok": True})
